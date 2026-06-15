"""Export pipeline results to Excel."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ..domain.data_types import PipelineResult

OUTPUT_DIR = Path(__file__).parent.parent.parent
FORMULA_PREFIXES = ("=", "+", "-", "@")


def escape_excel_text(value):
    """Force potentially executable spreadsheet text to remain plain text."""
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def export_result(result: PipelineResult, query_number: int) -> Path:
    """Export the current pipeline result to an xlsx file. Returns the file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"result_q{query_number}_{ts}.xlsx"
    path = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")

    def _hdr(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Summary table
    ws["A1"] = "Model"
    ws["B1"] = "Trucks"
    ws["C1"] = "Total Cost/month (USD)"
    ws["D1"] = "Total km/month"
    ws["E1"] = "Fixed Cost"
    ws["F1"] = "Variable Cost"
    ws["G1"] = "Overtime Cost"

    for cell in ws[1]:
        _hdr(cell)

    rows = []
    if result.lbl_result:
        r = result.lbl_result
        rows.append(["Lane-by-Lane", r.trucks, r.total_cost, r.total_km, r.fixed_cost, r.variable_cost, r.overtime_cost_total])
    if result.wct_result:
        r = result.wct_result
        rows.append(["Weighted Cycle Time", r.trucks, r.total_cost, r.total_km, r.fixed_cost, r.variable_cost, r.overtime_cost_total])
    mr = result.milp_result
    if mr.feasible:
        rows.append(["Solver (MILP)", mr.trucks, mr.total_cost, mr.total_km, mr.fixed_cost, mr.variable_cost, mr.overtime_cost_total])
    else:
        rows.append(["Solver (MILP)", "—", "No feasible solution", "—", "—", "—", "—"])

    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=escape_excel_text(val))

    # Scenario params
    ws2 = wb.create_sheet("Parameters")
    ws2["A1"] = "Parameter"
    ws2["B1"] = "Value"
    for cell in ws2[1]:
        _hdr(cell)

    p = result.scenario_params
    params_rows = [
        ("Payload (t)", p.payload),
        ("Speed loaded (km/h)", p.speed_loaded),
        ("Speed empty (km/h)", p.speed_empty),
        ("Availability", p.availability),
        ("Overtime (h/day)", p.overtime_hours),
        ("Variable cost/km (USD)", p.variable_cost_per_km),
        ("Fixed cost/truck/month (USD)", p.fixed_cost_per_truck_month),
        ("Working days/month", p.working_days),
        ("Net driving hours/day", p.net_driving_hours),
        *[(f"Terminal {tid} active", str(active)) for tid, active in p.terminals_active.items()],
        ("Objective", p.objective),
        ("Min coverage count", str(p.min_coverage_count)),
        ("Budget (USD)", str(p.budget)),
    ]
    for i, (k, v) in enumerate(params_rows, start=2):
        ws2.cell(row=i, column=1, value=escape_excel_text(k))
        ws2.cell(row=i, column=2, value=escape_excel_text(v))

    # MILP assignments
    if mr.feasible and mr.assignments:
        ws3 = wb.create_sheet("MILP Assignments")
        ws3["A1"] = "Collection Point"
        ws3["B1"] = "Terminal"
        for cell in ws3[1]:
            _hdr(cell)
        for i, (cp, t) in enumerate(sorted(mr.assignments.items()), start=2):
            ws3.cell(row=i, column=1, value=escape_excel_text(cp))
            ws3.cell(row=i, column=2, value=escape_excel_text(t))

    # Insight
    if result.insight:
        ws4 = wb.create_sheet("Insight")
        ws4["A1"] = "LLM Insight"
        ws4["A1"].font = Font(bold=True)
        ws4["A2"] = escape_excel_text(result.insight)
        ws4["A2"].alignment = Alignment(wrap_text=True)
        ws4.column_dimensions["A"].width = 80

    # Auto-width for summary
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(path)
    return path
